import json
import os

import openpyxl
import pandas
def create_folder_if_not_exists(file_path):
    # 获取文件夹路径
    folder_path = os.path.dirname(file_path)

    # 检查文件夹是否存在，如果不存在则创建
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"文件夹已创建: {folder_path}")
    else:
        print(f"文件夹已存在: {folder_path}")

def get_top_n(n, l, k_v, Best=True):
    lists = []
    # print(l)
    number_flag = 0
    min_value = 0
    num = 0
    for j in sorted(k_v.items(), key=lambda y: y[1], reverse=True):
        # print(j)
        if number_flag >= n:
            break
        else:
            number_flag = number_flag + 1
            min_value = j[1]
            lists.append(j[0])
    # print(lists)

    for i in lists:
        for j in l:
            if i == j:
                num += 1
    if not Best:
        for i in l:
            if i not in lists:
                # print('错误:', l)
                # print('lists: ', lists)
                # print('num: ', num)
                return False, num
            else:
                # print('错误:', l)
                # print('lists: ', lists)
                # print('num: ', num)
                return True, num

        for j in sorted(k_v.items(), key=lambda y: y[1], reverse=True):
            if j[1] == min_value:
                lists.append(j[0])
        for i in l:
            if k_v[i] != min_value:
                print('错误:', l)
                print('lists: ', lists)
                print('num: ', num)
                return True, num
            else:
                lists.remove(i)
                num = num - 1
        if len(lists) >= n:
            print('错误:', l)
            print('lists: ', lists)
            print('num: ', num)
            return False, num
        else:
            print('错误:', l)
            print('lists: ', lists)
            print('num: ', num)
            return True, num
    if Best:
        for j in sorted(k_v.items(), key=lambda y: y[1], reverse=True):
            if j[1] == min_value:
                lists.append(j[0])
        print(n, ' : ', lists)

        for i in l:
            if i in lists:
                return True
        return False


def key_value(list):
    k_v = {}
    for i in range(0, len(list)):
        k_v[i] = list[i]
    return k_v


def TOP_N(data_value):
    filepath = f'result/{data_value}/w_top_n.xlsx'
    create_folder_if_not_exists(filepath)
    f = openpyxl.Workbook()
    sheet1 = f.create_sheet()
    sheet1.title = 'top_1'
    sheet2 = f.create_sheet()
    sheet2.title = 'top_3'
    sheet3 = f.create_sheet()
    sheet3.title = 'top_5'

    row0 = [' ', 'Jaccard', 'Tarantula', 'Dstar', 'Hamann', 'Wong1', 'Hamming']
    for i in range(len(row0)):
        sheet1.cell(column=i + 1, row=1).value = row0[i]
        sheet2.cell(column=i + 1, row=1).value = row0[i]
        sheet3.cell(column=i + 1, row=1).value = row0[i]
    line0 = [' ', 'MBFL', 'MBFL_G', 'increase']

    for i in range(len(line0)):
        sheet1.cell(column=1, row=i + 1).value = line0[i]
        sheet2.cell(column=1, row=i + 1).value = line0[i]
        sheet3.cell(column=1, row=i + 1).value = line0[i]

    mbfl_GP13_top_1 = 0
    mbfl_Ochiai_top_1 = 0
    mbfl_Jaccard_top_1 = 0
    mbfl_OP2_top_1 = 0
    mbfl_Tarantula_top_1 = 0
    mbfl_Dstar_top_1 = 0
    mbfl_Dice_top_1 = 0
    mbfl_Hamann_top_1 = 0
    mbfl_Wong1_top_1 = 0
    mbfl_Hamming_top_1 = 0

    mbfl_GP13_top_3 = 0
    mbfl_Ochiai_top_3 = 0
    mbfl_Jaccard_top_3 = 0
    mbfl_OP2_top_3 = 0
    mbfl_Tarantula_top_3 = 0
    mbfl_Dstar_top_3 = 0
    mbfl_Dice_top_3 = 0
    mbfl_Hamann_top_3 = 0
    mbfl_Wong1_top_3 = 0
    mbfl_Hamming_top_3 = 0

    mbfl_GP13_top_5 = 0
    mbfl_Ochiai_top_5 = 0
    mbfl_Jaccard_top_5 = 0
    mbfl_OP2_top_5 = 0
    mbfl_Tarantula_top_5 = 0
    mbfl_Dstar_top_5 = 0
    mbfl_Dice_top_5 = 0
    mbfl_Hamann_top_5 = 0
    mbfl_Wong1_top_5 = 0
    mbfl_Hamming_top_5 = 0

    pr_mbfl_GP13_top_1 = 0
    pr_mbfl_Ochiai_top_1 = 0
    pr_mbfl_Jaccard_top_1 = 0
    pr_mbfl_OP2_top_1 = 0
    pr_mbfl_Tarantula_top_1 = 0
    pr_mbfl_Dstar_top_1 = 0
    pr_mbfl_Dice_top_1 = 0
    pr_mbfl_Hamann_top_1 = 0
    pr_mbfl_Wong1_top_1 = 0
    pr_mbfl_Hamming_top_1 = 0

    pr_mbfl_GP13_top_3 = 0
    pr_mbfl_Ochiai_top_3 = 0
    pr_mbfl_Jaccard_top_3 = 0
    pr_mbfl_OP2_top_3 = 0
    pr_mbfl_Tarantula_top_3 = 0
    pr_mbfl_Dstar_top_3 = 0
    pr_mbfl_Dice_top_3 = 0
    pr_mbfl_Hamann_top_3 = 0
    pr_mbfl_Wong1_top_3 = 0
    pr_mbfl_Hamming_top_3 = 0

    pr_mbfl_GP13_top_5 = 0
    pr_mbfl_Ochiai_top_5 = 0
    pr_mbfl_Jaccard_top_5 = 0
    pr_mbfl_OP2_top_5 = 0
    pr_mbfl_Tarantula_top_5 = 0
    pr_mbfl_Dstar_top_5 = 0
    pr_mbfl_Dice_top_5 = 0
    pr_mbfl_Hamann_top_5 = 0
    pr_mbfl_Wong1_top_5 = 0
    pr_mbfl_Hamming_top_5 = 0

    # print("GET JSON FILE FINISHED!", end='\n\n')
    # data = datas[i]

    i = 0
    totle_num = 0
    for data in datas:
        # i += 1
        # if i == 2: break
        data_name = data['proj']
        ans = data['ans']
        methods_num = len(ans)
        totle_num += methods_num
        print(data_name, "错误方法为：", ans)

        formulas = ['Jaccard', 'Tarantula', 'Dstar', 'Hamann', 'Wong1', 'Hamming']
        for formula in formulas:
            try:
                with open(f'{data_value}_SUS/MBFL/{data_name}{formula}_sus.txt', 'r') as mbfl_f:
                    mbfl_sus = key_value(eval(mbfl_f.readline()))
                    print(mbfl_sus)

                    # top_1
                    FLAG, n_num = get_top_n(1, ans, mbfl_sus, False)
                    if n_num != 0:
                        print('MBFL', n_num)
                    if formula == 'GP13':
                        if FLAG:
                            mbfl_GP13_top_1 += n_num
                    if formula == 'Ochiai':
                        if FLAG:
                            mbfl_Ochiai_top_1 += n_num
                    if formula == 'Jaccard':
                        if FLAG:
                            mbfl_Jaccard_top_1 += n_num
                    if formula == 'OP2':
                        if FLAG:
                            mbfl_OP2_top_1 += n_num
                    if formula == 'Tarantula':
                        if FLAG:
                            mbfl_Tarantula_top_1 += n_num
                    if formula == 'Dstar':
                        if FLAG:
                            mbfl_Dstar_top_1 += n_num
                    if formula == 'Dice':
                        if FLAG:
                            mbfl_Dice_top_1 += n_num
                    if formula == 'Hamann':
                        if FLAG:
                            mbfl_Hamann_top_1 += n_num
                    if formula == 'Wong1':
                        if FLAG:
                            mbfl_Wong1_top_1 += n_num
                    if formula == 'Hamming':
                        if FLAG:
                            mbfl_Hamming_top_1 += n_num
                    # top_3
                    FLAG, n_num = get_top_n(3, ans, mbfl_sus, False)
                    if formula == 'GP13':
                        if FLAG:
                            mbfl_GP13_top_3 += n_num
                    if formula == 'Ochiai':
                        if FLAG:
                            mbfl_Ochiai_top_3 += n_num
                    if formula == 'Jaccard':
                        if FLAG:
                            mbfl_Jaccard_top_3 += n_num
                    if formula == 'OP2':
                        if FLAG:
                            mbfl_OP2_top_3 += n_num
                    if formula == 'Tarantula':
                        if FLAG:
                            mbfl_Tarantula_top_3 += n_num
                    if formula == 'Dstar':
                        if FLAG:
                            mbfl_Dstar_top_3 += n_num
                    if formula == 'Dice':
                        if FLAG:
                            mbfl_Dice_top_3 += n_num
                    if formula == 'Hamann':
                        if FLAG:
                            mbfl_Hamann_top_3 += n_num
                    if formula == 'Wong1':
                        if FLAG:
                            mbfl_Wong1_top_3 += n_num
                    if formula == 'Hamming':
                        if FLAG:
                            mbfl_Hamming_top_3 += n_num
                    # top_5
                    FLAG, n_num = get_top_n(5, ans, mbfl_sus, False)
                    if formula == 'GP13':
                        if FLAG:
                            mbfl_GP13_top_5 += n_num
                    if formula == 'Ochiai':
                        if FLAG:
                            mbfl_Ochiai_top_5 += n_num
                    if formula == 'Jaccard':
                        if FLAG:
                            mbfl_Jaccard_top_5 += n_num
                    if formula == 'OP2':
                        if FLAG:
                            mbfl_OP2_top_5 += n_num
                    if formula == 'Tarantula':
                        if FLAG:
                            mbfl_Tarantula_top_5 += n_num
                    if formula == 'Dstar':
                        if FLAG:
                            mbfl_Dstar_top_5 += n_num
                    if formula == 'Dice':
                        if FLAG:
                            mbfl_Dice_top_5 += n_num
                    if formula == 'Hamann':
                        if FLAG:
                            mbfl_Hamann_top_5 += n_num
                    if formula == 'Wong1':
                        if FLAG:
                            mbfl_Wong1_top_5 += n_num
                    if formula == 'Hamming':
                        if FLAG:
                            mbfl_Hamming_top_5 += n_num
            except:
                continue

        for formula in formulas:
            try:
                with open(f'{data_value}_SUS/MBFL_G/{data_name}{formula}.txt', 'r') as mbfl_f:
                    pr_mbfl_sus = key_value(eval(mbfl_f.readline()))
                    print(pr_mbfl_sus)
                    # top_1
                    FLAG, n_num = get_top_n(1, ans, pr_mbfl_sus, False)
                    if n_num != 0:
                        print('PRMBFL', n_num)
                    if formula == 'GP13':
                        if FLAG:
                            pr_mbfl_GP13_top_1 += n_num
                    if formula == 'Ochiai':
                        if FLAG:
                            pr_mbfl_Ochiai_top_1 += n_num
                    if formula == 'Jaccard':
                        if FLAG:
                            pr_mbfl_Jaccard_top_1 += n_num
                    if formula == 'OP2':
                        if FLAG:
                            pr_mbfl_OP2_top_1 += n_num
                    if formula == 'Tarantula':
                        if FLAG:
                            pr_mbfl_Tarantula_top_1 += n_num
                    if formula == 'Dstar':
                        if FLAG:
                            pr_mbfl_Dstar_top_1 += n_num
                    if formula == 'Dice':
                        if FLAG:
                            pr_mbfl_Dice_top_1 += n_num
                    if formula == 'Hamann':
                        if FLAG:
                            pr_mbfl_Hamann_top_1 += n_num
                    if formula == 'Wong1':
                        if FLAG:
                            pr_mbfl_Wong1_top_1 += n_num
                    if formula == 'Hamming':
                        if FLAG:
                            pr_mbfl_Hamming_top_1 += n_num
                    # top_3
                    FLAG, n_num = get_top_n(3, ans, pr_mbfl_sus, False)
                    if formula == 'GP13':
                        if FLAG:
                            pr_mbfl_GP13_top_3 += n_num
                    if formula == 'Ochiai':
                        if FLAG:
                            pr_mbfl_Ochiai_top_3 += n_num
                    if formula == 'Jaccard':
                        if FLAG:
                            pr_mbfl_Jaccard_top_3 += n_num
                    if formula == 'OP2':
                        if FLAG:
                            pr_mbfl_OP2_top_3 += n_num
                    if formula == 'Tarantula':
                        if FLAG:
                            pr_mbfl_Tarantula_top_3 += n_num
                    if formula == 'Dstar':
                        if FLAG:
                            pr_mbfl_Dstar_top_3 += n_num
                    if formula == 'Dice':
                        if FLAG:
                            pr_mbfl_Dice_top_3 += n_num
                    if formula == 'Hamann':
                        if FLAG:
                            pr_mbfl_Hamann_top_3 += n_num
                    if formula == 'Wong1':
                        if FLAG:
                            pr_mbfl_Wong1_top_3 += n_num
                    if formula == 'Hamming':
                        if FLAG:
                            pr_mbfl_Hamming_top_3 += n_num

                    # top_5
                    FLAG, n_num = get_top_n(5, ans, pr_mbfl_sus, False)
                    if formula == 'GP13':
                        if FLAG:
                            pr_mbfl_GP13_top_5 += n_num
                    if formula == 'Ochiai':
                        if FLAG:
                            pr_mbfl_Ochiai_top_5 += n_num
                    if formula == 'Jaccard':
                        if FLAG:
                            pr_mbfl_Jaccard_top_5 += n_num
                    if formula == 'OP2':
                        if FLAG:
                            pr_mbfl_OP2_top_5 += n_num
                    if formula == 'Tarantula':
                        if FLAG:
                            pr_mbfl_Tarantula_top_5 += n_num
                    if formula == 'Dstar':
                        if FLAG:
                            pr_mbfl_Dstar_top_5 += n_num
                    if formula == 'Dice':
                        if FLAG:
                            pr_mbfl_Dice_top_5 += n_num
                    if formula == 'Hamann':
                        if FLAG:
                            pr_mbfl_Hamann_top_5 += n_num
                    if formula == 'Wong1':
                        if FLAG:
                            pr_mbfl_Wong1_top_5 += n_num
                    if formula == 'Hamming':
                        if FLAG:
                            pr_mbfl_Hamming_top_5 += n_num
            except:
                continue

    # print('   MBFL TOP1 ANS  ', '   mbfl_GP13_top_1 = ', mbfl_GP13_top_1, '   mbfl_Ochiai_top_1 = ', mbfl_Ochiai_top_1,
    #       '   mbfl_Jaccard_top_1 = ', mbfl_Jaccard_top_1, '   mbfl_OP2_top_1 = ', mbfl_OP2_top_1,
    #       '   mbfl_Tarantula_top_1 = ', mbfl_Tarantula_top_1)
    # print('pr_mbfl TOP1 ANS  ', 'pr_mbfl_GP13_top_1 = ', pr_mbfl_GP13_top_1, 'pr_mbfl_Ochiai_top_1 = ',
    #       pr_mbfl_Ochiai_top_1, 'pr_mbfl_Jaccard_top_1 = ', pr_mbfl_Jaccard_top_1, 'pr_mbfl_OP2_top_1 = ',
    #       pr_mbfl_OP2_top_1, 'pr_mbfl_Tarantula_top_1 = ', pr_mbfl_Tarantula_top_1)
    # print('------------------------------------')
    #
    # print('   MBFL TOP3 ANS  ', '   mbfl_GP13_top_3 = ', mbfl_GP13_top_3, '   mbfl_Ochiai_top_3 = ', mbfl_Ochiai_top_3,
    #       '   mbfl_Jaccard_top_3 = ', mbfl_Jaccard_top_3, '   mbfl_OP2_top_3 = ', mbfl_OP2_top_3,
    #       '   mbfl_Tarantula_top_3 = ', mbfl_Tarantula_top_3)
    # print('pr_mbfl TOP3 ANS  ', 'pr_mbfl_GP13_top_3 = ', pr_mbfl_GP13_top_3, 'pr_mbfl_Ochiai_top_3 = ',
    #       pr_mbfl_Ochiai_top_3, 'pr_mbfl_Jaccard_top_3 = ', pr_mbfl_Jaccard_top_3, 'pr_mbfl_OP2_top_3 = ',
    #       pr_mbfl_OP2_top_3, 'pr_mbfl_Tarantula_top_3 = ', pr_mbfl_Tarantula_top_3)
    # print('------------------------------------')
    #
    # print('   MBFL TOP5 ANS  ', '   mbfl_GP13_top_5 = ', mbfl_GP13_top_5, '   mbfl_Ochiai_top_5 = ', mbfl_Ochiai_top_5,
    #       '   mbfl_Jaccard_top_5 = ', mbfl_Jaccard_top_5, '   mbfl_OP2_top_5 = ', mbfl_OP2_top_5,
    #       '   mbfl_Tarantula_top_5 = ', mbfl_Tarantula_top_5)
    # print('pr_mbfl TOP5 ANS  ', 'pr_mbfl_GP13_top_5 = ', pr_mbfl_GP13_top_5, 'pr_mbfl_Ochiai_top_5 = ',
    #       pr_mbfl_Ochiai_top_5, 'pr_mbfl_Jaccard_top_5 = ', pr_mbfl_Jaccard_top_5, 'pr_mbfl_OP2_top_5 = ',
    #       pr_mbfl_OP2_top_5, 'pr_mbfl_Tarantula_top_5 = ', pr_mbfl_Tarantula_top_5)
    #
    # print('------------------------------------')
    data1 = [
        [mbfl_Jaccard_top_1, pr_mbfl_Jaccard_top_1, pr_mbfl_Jaccard_top_1 - mbfl_Jaccard_top_1],
        [mbfl_Tarantula_top_1, pr_mbfl_Tarantula_top_1, pr_mbfl_Tarantula_top_1 - mbfl_Tarantula_top_1],
        [mbfl_Dstar_top_1, pr_mbfl_Dstar_top_1, pr_mbfl_Dstar_top_1 - mbfl_Dstar_top_1],
        [mbfl_Hamann_top_1, pr_mbfl_Hamann_top_1, pr_mbfl_Hamann_top_1 - mbfl_Hamann_top_1],
        [mbfl_Wong1_top_1, pr_mbfl_Wong1_top_1, pr_mbfl_Wong1_top_1 - mbfl_Wong1_top_1],
        [mbfl_Hamming_top_1, pr_mbfl_Hamming_top_1, pr_mbfl_Hamming_top_1 - mbfl_Hamming_top_1],
    ]
    data2 = [
        [mbfl_Jaccard_top_3, pr_mbfl_Jaccard_top_3, pr_mbfl_Jaccard_top_3 - mbfl_Jaccard_top_3],
        [mbfl_Tarantula_top_3, pr_mbfl_Tarantula_top_3, pr_mbfl_Tarantula_top_3 - mbfl_Tarantula_top_3],
        [mbfl_Dstar_top_3, pr_mbfl_Dstar_top_3, pr_mbfl_Dstar_top_3 - mbfl_Dstar_top_3],
        [mbfl_Hamann_top_3, pr_mbfl_Hamann_top_3, pr_mbfl_Hamann_top_3 - mbfl_Hamann_top_3],
        [mbfl_Wong1_top_3, pr_mbfl_Wong1_top_3, pr_mbfl_Wong1_top_3 - mbfl_Wong1_top_3],
        [mbfl_Hamming_top_3, pr_mbfl_Hamming_top_3, pr_mbfl_Hamming_top_3 - mbfl_Hamming_top_3],

    ]
    data3 = [
        [mbfl_Jaccard_top_5, pr_mbfl_Jaccard_top_5, pr_mbfl_Jaccard_top_5 - mbfl_Jaccard_top_5],
        [mbfl_Tarantula_top_5, pr_mbfl_Tarantula_top_5, pr_mbfl_Tarantula_top_5 - mbfl_Tarantula_top_5],
        [mbfl_Dstar_top_5, pr_mbfl_Dstar_top_5, pr_mbfl_Dstar_top_5 - mbfl_Dstar_top_5],
        [mbfl_Hamann_top_5, pr_mbfl_Hamann_top_5, pr_mbfl_Hamann_top_5 - mbfl_Hamann_top_5],
        [mbfl_Wong1_top_5, pr_mbfl_Wong1_top_5, pr_mbfl_Wong1_top_5 - mbfl_Wong1_top_5],
        [mbfl_Hamming_top_5, pr_mbfl_Hamming_top_5, pr_mbfl_Hamming_top_5 - mbfl_Hamming_top_5],

    ]
    for r in range(len(data1)):
        for col in range(len(data1[0])):
            try:
                sheet1.cell(column=r + 2, row=col + 2).value = data1[r][col]
            except:
                print(r - 2, col - 2)
    for r in range(len(data2)):
        for col in range(len(data2[0])):
            try:
                sheet2.cell(column=r + 2, row=col + 2).value = data2[r][col]
            except:
                print(r - 2, col - 2)
    for r in range(len(data3)):
        for col in range(len(data3[0])):
            try:
                sheet3.cell(column=r + 2, row=col + 2).value = data3[r][col]
            except:
                print(r - 2, col - 2)
    sheet1.cell(column=5, row=5).value = totle_num
    sheet2.cell(column=5, row=5).value = totle_num
    sheet3.cell(column=5, row=5).value = totle_num
    f.save(filepath)


def get_rank(location_list, k_v):
    rank = 9999999

    for location in location_list:
        this_rank = 1
        sus = k_v[location]
        for i in k_v:
            if k_v[i] > sus:
                this_rank += 1
        if this_rank < rank:
            rank = this_rank

    return rank


def average_rank(location_list, k_v):
    rank = 9999999

    for location in location_list:
        this_rank = 1
        sus = k_v[location]
        for i in k_v:
            if k_v[i] > sus:
                this_rank += 1
            if k_v[i] == sus:
                this_rank += 0.5
        if this_rank < rank:
            rank = this_rank

    return rank


def EXAM(data_value, a, d):
    filepath = f'result/{data_value}/EXAM.xlsx'
    f = openpyxl.Workbook()

    formulas = [' ', 'Jaccard', 'Tarantula', 'OP2', 'GP13', 'Ochiai', 'Dstar', 'Dice', 'Hamann', 'Wong1', 'Hamming']

    for formula in formulas:
        ok = 0
        mbfl_avg = 0
        pr_avg = 0
        sheet = f.create_sheet()
        sheet.title = formula
        sheet.cell(column=1, row=1).value = formula
        sheet.cell(column=1, row=2).value = 'MBFL'
        sheet.cell(column=1, row=3).value = 'PRMBFL'
        better, bad = 0, 0

        i = 0
        for data in datas:

            data_name = data['proj']
            ans = data['ans']
            methods_num = len(data['methods'])
            # print(data_name, "错误方法为：", ans)
            sheet.cell(column=i + 2, row=1).value = data_name

            try:
                with open(f'{data_value}_SUS/MBFL/{data_name}{formula}_sus.txt', 'r') as mbfl_f:
                    mbfl_sus = key_value(eval(mbfl_f.readline()))
                    rank = average_rank(ans, mbfl_sus)
                    sheet.cell(column=i + 2, row=2).value = rank / methods_num
                    mbfl_avg += rank / methods_num
                    # print(mbfl_sus)
                    # print(rank )
                with open(f'{data_value}_SUS/PR_MBFL/{data_name}{formula}.txt', 'r') as mbfl_f:
                    pr_mbfl_sus = key_value(eval(mbfl_f.readline()))
                    rank2 = average_rank(ans, pr_mbfl_sus)
                    sheet.cell(column=i + 2, row=3).value = rank2 / methods_num
                    pr_avg += rank2 / methods_num

            except:
                continue
                # print(rank2 / methods_num)
            # print(formula, ' : ', rank/ methods_num , '  ----> ',rank2/ methods_num)

            if rank / methods_num - rank2 / methods_num > 0:
                better += 1
            elif rank / methods_num - rank2 / methods_num < 0:
                bad += 1
            sheet.cell(column=i + 2, row=4).value = rank / methods_num - rank2 / methods_num
            sheet.cell(column=3, row=4)
            if rank2 <= rank:
                ok += 1
            i = i + 1
        sheet.cell(column=3, row=5).value = "better" + str(better)
        sheet.cell(column=4, row=5).value = 'bad' + str(bad)

        print(mbfl_avg, pr_avg)
    f.save(filepath)
    # print(pandas.read_csv(f))

    return


def MFR(data_value, a, d):
    filepath = f'result/{data_value}/MFR.xlsx'
    f = openpyxl.Workbook()

    formulas = [' ', 'Jaccard', 'Tarantula', 'OP2', 'GP13', 'Ochiai', 'Dstar', 'Dice', 'Hamann', 'Wong1', 'Hamming']

    for formula in formulas:
        ok = 0

        sheet = f.create_sheet()
        sheet.title = formula
        sheet.cell(column=1, row=1).value = formula
        sheet.cell(column=1, row=2).value = 'MBFL'
        sheet.cell(column=1, row=3).value = 'PRMBFL'
        better, bad = 0, 0

        i = 0
        for data in datas:
            try:
                data_name = data['proj']
                ans = data['ans']
                methods_num = len(data['methods'])
                print(data_name, "错误方法为：", ans)

                sheet.cell(column=i + 2, row=1).value = data_name

                with open(f'{data_value}_SUS/MBFL/{data_name}{formula}_sus.txt', 'r') as mbfl_f:
                    mbfl_sus = key_value(eval(mbfl_f.readline()))
                    rank = get_rank(ans, mbfl_sus)
                    sheet.cell(column=i + 2, row=2).value = rank
                    # print(mbfl_sus)
                    # print(rank )
                with open(f'{data_value}_SUS/PR_MBFL/{data_name}{formula}.txt', 'r') as mbfl_f:
                    pr_mbfl_sus = key_value(eval(mbfl_f.readline()))
                    rank2 = get_rank(ans, pr_mbfl_sus)
                    sheet.cell(column=i + 2, row=3).value = rank2

                    # print(rank2 / methods_num)
                print(formula, ' : ', rank / methods_num, '  ----> ', rank2 / methods_num)

                if rank - rank2 > 0:
                    better += 1
                elif rank - rank2 < 0:
                    bad += 1
                sheet.cell(column=i + 2, row=4).value = rank - rank2
                sheet.cell(column=3, row=4)
                if rank2 <= rank:
                    ok += 1
                i = i + 1
            except:
                continue
        sheet.cell(column=3, row=5).value = "better" + str(better)
        sheet.cell(column=4, row=5).value = 'bad' + str(bad)

        print(ok)
    f.save(filepath)

    return


def MAR(data_value, a, d):
    filepath = f'result/{data_value}/MAR.xlsx'
    f = openpyxl.Workbook()

    formulas = [ 'Jaccard', 'Tarantula', 'OP2', 'GP13', 'Ochiai', 'Dstar', 'Dice', 'Hamann', 'Wong1', 'Hamming']

    for formula in formulas:
        ok = 0

        sheet = f.create_sheet()
        sheet.title = formula
        sheet.cell(column=1, row=1).value = formula
        sheet.cell(column=1, row=2).value = 'MBFL'
        sheet.cell(column=1, row=3).value = 'PRMBFL'
        better, bad = 0, 0

        i = 0
        for data in datas:
            data_name = data['proj']
            ans = data['ans']
            methods_num = len(data['methods'])
            print(data_name, "错误方法为：", ans)
            sheet.cell(column=i + 2, row=1).value = data_name

            try:
                with open(f'{data_value}_SUS/MBFL/{data_name}{formula}_sus.txt', 'r') as mbfl_f:
                    mbfl_sus = key_value(eval(mbfl_f.readline()))
                    rank = average_rank(ans, mbfl_sus)
                    sheet.cell(column=i + 2, row=2).value = rank
                    # print(mbfl_sus)
                    # print(rank )
                with open(f'{data_value}_SUS/PR_MBFL/{data_name}{formula}.txt', 'r') as mbfl_f:
                    pr_mbfl_sus = key_value(eval(mbfl_f.readline()))
                    rank2 = average_rank(ans, pr_mbfl_sus)
                    sheet.cell(column=i + 2, row=3).value = rank2

                    # print(rank2 / methods_num)
            except:
                continue
            print(formula, ' : ', rank / methods_num, '  ----> ', rank2 / methods_num)

            if rank - rank2 > 0:
                better += 1
            elif rank - rank2 < 0:
                bad += 1
            sheet.cell(column=i + 2, row=4).value = rank - rank2
            sheet.cell(column=3, row=4)
            if rank2 < rank:
                ok += 1
            i = i + 1
        sheet.cell(column=3, row=5).value = "better" + str(better)
        sheet.cell(column=4, row=5).value = 'bad' + str(bad)

        print(ok)
    f.save(filepath)

    return


if __name__ == '__main__':

    cal_names = ['Dstar']
    dataset = ['tcas']
    for data_value in dataset:
        with open(f'Data/{data_value}.json', 'r') as rf:
            datas = json.load(rf)
        # for a in a_list:
        #     for d in d_list:。。
        TOP_N(data_value)

